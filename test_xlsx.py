from openpyxl import load_workbook
import os
# TODO оформить в тест, добавить ассерты и использовать универсальный путь
file_path = os.path.join('resources/file_example_XLSX_50.xlsx')
def test_xlsx():
    workbook = load_workbook(file_path)
    sheet = workbook.active
    print("Пересечение: ", sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == "Mara"